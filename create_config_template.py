#!/usr/bin/env python3
"""
Erstellt eine neue Analysis-Config.xlsx Vorlage mit Filter-Spalten.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_config_template(output_file="Analysis-Config_v1.3.xlsx"):
    """Erstellt eine Excel-Konfigurationsvorlage mit Filter-Spalten."""
    
    # Arbeitsmappe erstellen
    wb = openpyxl.Workbook()
    
    # =============================================================================
    # SHEET 1: VARIABLEN
    # =============================================================================
    
    ws_variablen = wb.active
    ws_variablen.title = "Variablen"
    
    # Spaltenüberschriften mit Formatierung
    headers = [
        "variable_name", "question_text", "data_type", "coding", 
        "min_value", "max_value", "reverse_coding", "use_NA", "filter"
    ]
    
    # Formatierung für Kopfzeile
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Kopfzeile schreiben
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_variablen.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Beispiel-Daten für Variablen-Sheet
    examples = [
        # variable_name, question_text, data_type, coding, min_value, max_value, reverse_coding, use_NA, filter
        ["SD01", "Geschlecht", "nominal_coded", "1=Weiblich;2=Männlich;3=Divers", "", "", "FALSE", "FALSE", ""],
        ["SD02", "Alter in Jahren", "numeric", "", "18", "99", "FALSE", "TRUE", "SD02 >= 18"],
        ["SD03", "Bildungsabschluss", "nominal_coded", "1=Hauptschule;2=Realschule;3=Abitur;4=Universität", "", "", "FALSE", "FALSE", ""],
        ["GP01", "Allgemeine Zufriedenheit", "ordinal", "1=Sehr unzufrieden;2=Unzufrieden;3=Neutral;4=Zufrieden;5=Sehr zufrieden", "1", "5", "FALSE", "FALSE", "SD01 == 1"],
        ["AS01", "Motivation für das Studium", "ordinal", "1=Sehr niedrig;2=Niedrig;3=Mittel;4=Hoch;5=Sehr hoch", "1", "5", "FALSE", "FALSE", ""],
        ["ZS01", "Zufriedenheit mit einzelnen Aspekten", "matrix", "1=Sehr unzufrieden;2=Unzufrieden;3=Neutral;4=Zufrieden;5=Sehr zufrieden", "1", "5", "FALSE", "FALSE", ""],
        ["NW01", "Welche Netzwerke nutzen Sie?", "matrix", "1=Ausgewählt", "", "", "FALSE", "FALSE", "SD02 >= 25"],
        ["zufriedenheit_index", "Zufriedenheits-Index (Mittelwert)", "numeric", "", "1", "5", "FALSE", "TRUE", "!is.na(zufriedenheit_index)"]
    ]
    
    # Beispiel-Daten schreiben
    for row_idx, example in enumerate(examples, start=2):
        for col_idx, value in enumerate(example, start=1):
            ws_variablen.cell(row=row_idx, column=col_idx, value=value)
    
    # Spaltenbreiten anpassen
    column_widths = [20, 40, 15, 40, 12, 12, 15, 12, 30]
    for i, width in enumerate(column_widths, start=1):
        ws_variablen.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # =============================================================================
    # SHEET 2: KREUZTABELLEN
    # =============================================================================
    
    ws_kreuztabellen = wb.create_sheet(title="Kreuztabellen")
    
    # Spaltenüberschriften
    headers = ["analysis_name", "variable_1", "variable_2", "statistical_test", "filter"]
    
    # Kopfzeile schreiben
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_kreuztabellen.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Beispiel-Daten für Kreuztabellen
    examples = [
        # analysis_name, variable_1, variable_2, statistical_test, filter
        ["Geschlecht_x_Zufriedenheit", "SD01", "GP01", "chi_square", ""],
        ["Alter_x_Motivation", "SD02", "AS01", "correlation", "SD02 >= 18"],
        ["Bildung_x_Zufriedenheit", "SD03", "zufriedenheit_index", "anova", "SD01 == 1"],
        ["Matrix_Zufriedenheit_x_Geschlecht", "ZS01", "SD01", "mann_whitney", ""],
        ["Netzwerk_x_Alter", "NW01", "SD02", "chi_square", "SD02 >= 25 & SD02 <= 60"]
    ]
    
    # Beispiel-Daten schreiben
    for row_idx, example in enumerate(examples, start=2):
        for col_idx, value in enumerate(example, start=1):
            ws_kreuztabellen.cell(row=row_idx, column=col_idx, value=value)
    
    # Spaltenbreiten anpassen
    column_widths = [30, 15, 15, 20, 40]
    for i, width in enumerate(column_widths, start=1):
        ws_kreuztabellen.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # =============================================================================
    # SHEET 3: REGRESSIONEN
    # =============================================================================
    
    ws_regressionen = wb.create_sheet(title="Regressionen")
    
    # Spaltenüberschriften
    headers = ["regression_name", "dependent_var", "independent_vars", "regression_type", "filter"]
    
    # Kopfzeile schreiben
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_regressionen.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Beispiel-Daten für Regressionen
    examples = [
        # regression_name, dependent_var, independent_vars, regression_type, filter
        ["Zufriedenheit_Modell", "zufriedenheit_index", "SD01;SD02;SD03", "linear", ""],
        ["Motivation_Modell", "AS01", "SD01;SD02;SD03", "linear", "SD02 >= 18"],
        ["Geschlecht_Regression", "GP01", "SD01", "t_test", ""],
        ["Interaktion_Modell", "zufriedenheit_index", "SD01*SD02;SD03", "linear", "!is.na(zufriedenheit_index)"],
        ["Mehrebenen_Modell", "GP01", "SD02;AS01", "multilevel", "SD01 == 1"]
    ]
    
    # Beispiel-Daten schreiben
    for row_idx, example in enumerate(examples, start=2):
        for col_idx, value in enumerate(example, start=1):
            ws_regressionen.cell(row=row_idx, column=col_idx, value=value)
    
    # Spaltenbreiten anpassen
    column_widths = [25, 20, 30, 20, 40]
    for i, width in enumerate(column_widths, start=1):
        ws_regressionen.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # =============================================================================
    # SHEET 4: TEXTANTWORTEN
    # =============================================================================
    
    ws_textantworten = wb.create_sheet(title="Textantworten")
    
    # Spaltenüberschriften
    headers = ["analysis_name", "text_variable", "sort_variable", "min_length", "include_empty", "filter"]
    
    # Kopfzeile schreiben
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_textantworten.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Beispiel-Daten für Textantworten
    examples = [
        # analysis_name, text_variable, sort_variable, min_length, include_empty, filter
        ["Verbesserungsvorschläge", "GP05[other]", "SD01", "3", "FALSE", ""],
        ["Sonstige_Bemerkungen", "AS08[other]", "SD03", "5", "TRUE", "SD02 >= 25"],
        ["Freitext_Feedback", "ZF01[other]", "", "10", "FALSE", ""],
        ["Kommentare_zum_Studium", "ST09[other]", "GP01", "3", "FALSE", "SD01 == 1"]
    ]
    
    # Beispiel-Daten schreiben
    for row_idx, example in enumerate(examples, start=2):
        for col_idx, value in enumerate(example, start=1):
            ws_textantworten.cell(row=row_idx, column=col_idx, value=value)
    
    # Spaltenbreiten anpassen
    column_widths = [25, 25, 20, 12, 15, 40]
    for i, width in enumerate(column_widths, start=1):
        ws_textantworten.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # =============================================================================
    # INFORMATIONSSHEET MIT FILTER-SYNTAX
    # =============================================================================
    
    ws_info = wb.create_sheet(title="Filter_Syntax_Hilfe")
    
    # Titel
    ws_info.cell(row=1, column=1, value="FILTER-SYNTAX FÜR DIE KONFIGURATION").font = Font(bold=True, size=14)
    ws_info.merge_cells('A1:F1')
    
    # Allgemeine Informationen
    info_rows = [
        ["", "", "", "", "", ""],
        ["Die Filter-Spalte erlaubt individuelle Filter für jede Variable/Analyse.", "", "", "", "", ""],
        ["", "", "", "", "", ""],
        ["Syntax-Beispiele:", "", "", "", "", ""],
        ["Einfache Vergleiche:", "SD01 == 1", "Variable SD01 muss Wert 1 haben", "", "", ""],
        ["", "ALTER >= 18", "Alter muss mindestens 18 sein", "", "", ""],
        ["Textvergleiche:", 'geschlecht == "weiblich"', "Geschlecht muss 'weiblich' sein", "", "", ""],
        ["", 'bildung != "kein Abschluss"', "Bildung darf nicht 'kein Abschluss' sein", "", "", ""],
        ["Logische Verknüpfungen:", "(SD01 == 1 & ALTER >= 25) | SD03 == 'hoch'", "Geschlecht=1 UND Alter≥25 ODER Bildung='hoch'", "", "", ""],
        ["Funktionen:", "!is.na(ZUFRIEDENHEIT) & ZUFRIEDENHEIT > 3", "Zufriedenheit nicht fehlend UND >3", "", "", ""],
        ["Matrix-Variablen:", "ZS01.001. == 1", "Matrix-Item ZS01[001] muss Wert 1 haben", "", "", ""],
        ["", "ZS01[001] == 1", "Alternative Schreibweise mit eckigen Klammern", "", "", ""],
        ["Komplexe Ausdrücke:", "(SD01 == 1 & SD02 >= 25) | (SD01 == 2 & SD02 <= 30)", "Komplexe logische Bedingungen", "", "", ""],
        ["", "", "", "", "", ""],
        ["WICHTIGE HINWEISE:", "", "", "", "", ""],
        ["• Filter werden für jede Variable/Analyse INDIVIDUELL angewendet", "", "", "", "", ""],
        ["• Filter-Ausdrücke müssen gültige R-Syntax verwenden", "", "", "", "", ""],
        ["• Variablennamen müssen im Datensatz vorhanden sein", "", "", "", "", ""],
        ["• Leere Filter-Zellen bedeuten: KEIN FILTER angewendet", "", "", "", "", ""],
        ["• Filter reduzieren die Fallzahl N für die jeweilige Analyse", "", "", "", "", ""],
        ["• Filter-Info wird im Excel-Output dokumentiert", "", "", "", "", ""],
    ]
    
    for row_idx, row_data in enumerate(info_rows, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            if value:
                cell = ws_info.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14)
                elif "Syntax-Beispiele:" in str(value) or "WICHTIGE HINWEISE:" in str(value):
                    cell.font = Font(bold=True)
    
    # Spaltenbreiten anpassen
    for i in range(1, 7):
        ws_info.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 25
    
    # Arbeitsmappe speichern
    wb.save(output_file)
    print(f"ERFOLG: Konfigurationsvorlage erstellt: {output_file}")
    print("\nSheets in der neuen Datei:")
    print("1. 'Variablen' - Variablendefinitionen mit Filter-Spalte")
    print("2. 'Kreuztabellen' - Kreuztabellen-Konfiguration mit Filter-Spalte")
    print("3. 'Regressionen' - Regressionsmodelle mit Filter-Spalte")
    print("4. 'Textantworten' - Textanalyse-Konfiguration mit Filter-Spalte")
    print("5. 'Filter_Syntax_Hilfe' - Dokumentation der Filter-Syntax")
    print("\nBeispiel-Filter in den Dummy-Daten:")
    print("- SD02 >= 18 (Alter mindestens 18 Jahre)")
    print("- SD01 == 1 (Nur weibliche Teilnehmer)")
    print("- !is.na(zufriedenheit_index) (Nur vollständige Indexwerte)")
    print("- SD02 >= 25 & SD02 <= 60 (Alter zwischen 25 und 60)")

if __name__ == "__main__":
    create_config_template()